#!/usr/bin/env python3
"""Generate cpt-tournament-bracket.xlsx from the tournament bracket data."""

from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, numbers
)
from openpyxl.utils import get_column_letter

wb = Workbook()

# ── Colour palette ─────────────────────────────────────────────────────────────
HDR_BG   = "1F4E79"   # dark blue
HDR_FG   = "FFFFFF"
ALT_BG   = "D6E4F0"   # light blue stripe
SIT_BG   = "D9D9D9"   # grey for sit-out cells
PLAY_BG  = "E2EFDA"   # light green for play cells
TOT_BG   = "FFF2CC"   # yellow for totals column

def header_style(cell, *, bold=True):
    cell.font      = Font(bold=bold, color=HDR_FG)
    cell.fill      = PatternFill("solid", fgColor=HDR_BG)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

def alt_fill(row_idx):
    return PatternFill("solid", fgColor=ALT_BG) if row_idx % 2 == 0 else None

def border():
    thin = Side(style="thin", color="AAAAAA")
    return Border(left=thin, right=thin, top=thin, bottom=thin)

def apply_border(cell):
    cell.border = border()

def set_col_width(ws, col, width):
    ws.column_dimensions[get_column_letter(col)].width = width


# ══════════════════════════════════════════════════════════════════════════════
# Sheet 1 — Schedule
# ══════════════════════════════════════════════════════════════════════════════
ws1 = wb.active
ws1.title = "Schedule"

ws1.row_dimensions[1].height = 30
schedule_data = [
    ("Round", "Teams in the ring", "Sitting out"),
    ("R1",  "1, 2, 3, 4",    "5, 7, 9, 10"),
    ("R2",  "5, 7, 9, 10",   "1, 2, 3, 4"),
    ("R3",  "1, 2, 5, 7",    "3, 4, 9, 10"),
    ("R4",  "3, 4, 9, 10",   "1, 2, 5, 7"),
    ("R5",  "1, 2, 9, 10",   "3, 4, 5, 7"),
    ("R6",  "3, 4, 5, 7",    "1, 2, 9, 10"),
    ("R7",  "1, 3, 5, 9",    "2, 4, 7, 10"),
    ("R8",  "2, 4, 7, 10",   "1, 3, 5, 9"),
    ("R9",  "1, 3, 7, 10",   "2, 4, 5, 9"),
    ("R10", "2, 4, 5, 9",    "1, 3, 7, 10"),
    ("R11", "1, 4, 5, 10",   "2, 3, 7, 9"),
    ("R12", "2, 3, 7, 9",    "1, 4, 5, 10"),
]
for r_idx, row in enumerate(schedule_data, 1):
    for c_idx, val in enumerate(row, 1):
        cell = ws1.cell(row=r_idx, column=c_idx, value=val)
        apply_border(cell)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        if r_idx == 1:
            header_style(cell)
        else:
            fill = alt_fill(r_idx)
            if fill:
                cell.fill = fill

set_col_width(ws1, 1, 10)
set_col_width(ws1, 2, 24)
set_col_width(ws1, 3, 24)


# ══════════════════════════════════════════════════════════════════════════════
# Sheet 2 — Team Schedule
# ══════════════════════════════════════════════════════════════════════════════
ws2 = wb.create_sheet("Team Schedule")

ws2.row_dimensions[1].height = 30
team_schedule = [
    ("Team", "Plays",                     "Rests"),
    ("1",    "R1 R3 R5 R7 R9 R11",        "R2 R4 R6 R8 R10 R12"),
    ("2",    "R1 R3 R5 R8 R10 R12",       "R2 R4 R6 R7 R9 R11"),
    ("3",    "R1 R4 R6 R7 R9 R12",        "R2 R3 R5 R8 R10 R11"),
    ("4",    "R1 R4 R6 R8 R10 R11",       "R2 R3 R5 R7 R9 R12"),
    ("5",    "R2 R3 R6 R7 R10 R11",       "R1 R4 R5 R8 R9 R12"),
    ("7",    "R2 R3 R6 R8 R9 R12",        "R1 R4 R5 R7 R10 R11"),
    ("9",    "R2 R4 R5 R7 R10 R12",       "R1 R3 R6 R8 R9 R11"),
    ("10",   "R2 R4 R5 R8 R9 R11",        "R1 R3 R6 R7 R10 R12"),
]
note = "Each team plays exactly 6 rounds. Every pair of teams meets exactly 2 or 3 times."

for r_idx, row in enumerate(team_schedule, 1):
    for c_idx, val in enumerate(row, 1):
        cell = ws2.cell(row=r_idx, column=c_idx, value=val)
        apply_border(cell)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        if r_idx == 1:
            header_style(cell)
        else:
            fill = alt_fill(r_idx)
            if fill:
                cell.fill = fill

note_cell = ws2.cell(row=len(team_schedule) + 2, column=1, value=note)
note_cell.font = Font(italic=True, color="555555")
ws2.merge_cells(
    start_row=len(team_schedule) + 2, start_column=1,
    end_row=len(team_schedule) + 2, end_column=3
)

set_col_width(ws2, 1, 10)
set_col_width(ws2, 2, 36)
set_col_width(ws2, 3, 36)


# ══════════════════════════════════════════════════════════════════════════════
# Sheet 3 — Scoring
# ══════════════════════════════════════════════════════════════════════════════
ws3 = wb.create_sheet("Scoring")

ws3.row_dimensions[1].height = 30
scoring_data = [
    ("Finish / Event",              "Points"),
    ("Last robot standing",         3),
    ("Eliminated 3rd (2nd-to-last)", 2),
    ("Eliminated 2nd",              1),
    ("Eliminated 1st",              0),
    ("Causing an elimination",      "+2 per robot eliminated"),
]
for r_idx, row in enumerate(scoring_data, 1):
    for c_idx, val in enumerate(row, 1):
        cell = ws3.cell(row=r_idx, column=c_idx, value=val)
        apply_border(cell)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        if r_idx == 1:
            header_style(cell)
        else:
            fill = alt_fill(r_idx)
            if fill:
                cell.fill = fill

set_col_width(ws3, 1, 34)
set_col_width(ws3, 2, 28)


# ══════════════════════════════════════════════════════════════════════════════
# Sheet 4 — Scoresheet
# ══════════════════════════════════════════════════════════════════════════════
ws4 = wb.create_sheet("Scoresheet")

teams   = [1, 2, 3, 4, 5, 7, 9, 10]
rounds  = list(range(1, 13))  # R1 … R12

# sit_out[round] = set of teams sitting out that round
sit_out = {
    1:  {5, 7, 9, 10},
    2:  {1, 2, 3, 4},
    3:  {3, 4, 9, 10},
    4:  {1, 2, 5, 7},
    5:  {3, 4, 5, 7},
    6:  {1, 2, 9, 10},
    7:  {2, 4, 7, 10},
    8:  {1, 3, 5, 9},
    9:  {2, 4, 5, 9},
    10: {1, 3, 7, 10},
    11: {2, 3, 7, 9},
    12: {1, 4, 5, 10},
}

# Header row
ws4.row_dimensions[1].height = 30
header_vals = ["Team"] + [f"R{r}" for r in rounds] + ["Total"]
for c_idx, val in enumerate(header_vals, 1):
    cell = ws4.cell(row=1, column=c_idx, value=val)
    apply_border(cell)
    cell.alignment = Alignment(horizontal="center", vertical="center")
    if val == "Total":
        cell.fill = PatternFill("solid", fgColor=TOT_BG)
        cell.font = Font(bold=True, color="000000")
    else:
        header_style(cell)

# Data rows
for r_idx, team in enumerate(teams, 2):
    fill = alt_fill(r_idx)

    # Team label
    team_cell = ws4.cell(row=r_idx, column=1, value=str(team))
    apply_border(team_cell)
    team_cell.alignment = Alignment(horizontal="center", vertical="center")
    team_cell.font = Font(bold=True)
    if fill:
        team_cell.fill = fill

    # Round columns (cols 2–13)
    score_cols = []
    for c_idx, rnd in enumerate(rounds, 2):
        cell = ws4.cell(row=r_idx, column=c_idx)
        apply_border(cell)
        cell.alignment = Alignment(horizontal="center", vertical="center")

        if team in sit_out[rnd]:
            cell.value = "—"
            cell.fill  = PatternFill("solid", fgColor=SIT_BG)
            cell.font  = Font(color="888888")
        else:
            cell.fill = PatternFill("solid", fgColor=PLAY_BG)
            score_cols.append(get_column_letter(c_idx))

    # Total column (col 14) — SUM of the green (play) cells
    total_cell = ws4.cell(row=r_idx, column=14)
    apply_border(total_cell)
    total_cell.alignment = Alignment(horizontal="center", vertical="center")
    total_cell.font = Font(bold=True)
    total_cell.fill = PatternFill("solid", fgColor=TOT_BG)
    if score_cols:
        refs = "+".join(f"{col}{r_idx}" for col in score_cols)
        total_cell.value = f"={refs}"

# Column widths
set_col_width(ws4, 1, 8)
for c in range(2, 14):
    set_col_width(ws4, c, 6)
set_col_width(ws4, 14, 10)

# Legend below the table
legend_row = len(teams) + 3
ws4.cell(row=legend_row,     column=1, value="Legend")
ws4.cell(row=legend_row,     column=1).font = Font(bold=True)
ws4.cell(row=legend_row + 1, column=1, value="Green cell = team plays this round (enter score)")
ws4.cell(row=legend_row + 1, column=1).fill = PatternFill("solid", fgColor=PLAY_BG)
ws4.cell(row=legend_row + 2, column=1, value="Grey  cell = team sits out this round")
ws4.cell(row=legend_row + 2, column=1).fill = PatternFill("solid", fgColor=SIT_BG)
ws4.merge_cells(start_row=legend_row + 1, start_column=1, end_row=legend_row + 1, end_column=5)
ws4.merge_cells(start_row=legend_row + 2, start_column=1, end_row=legend_row + 2, end_column=5)
for offset in (1, 2):
    ws4.cell(row=legend_row + offset, column=1).alignment = Alignment(horizontal="left")
    ws4.cell(row=legend_row + offset, column=1).font = Font(italic=True, size=9)

# ── Write file ──────────────────────────────────────────────────────────────
out_path = "/Users/davecheng/Documents/Code/Class-Notes-and-Addenda/TEJ4M/unit-4-rpi/cpt-tournament-bracket.xlsx"
wb.save(out_path)
print(f"Saved: {out_path}")
